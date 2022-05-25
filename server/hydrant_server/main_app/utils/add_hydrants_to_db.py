import openpyxl
from ..models import Hydrant, Profile



def parser_address(address):
    array_address = address.split(' ')
    if array_address[0] == 'г.':
        return {'region': 'г. Минск',
                'district': f'{array_address[2]} район',
                'village_council': '',
                'town': 'г. Минск'}
    elif array_address[5] == 'г.':
        return {'region': f'{array_address[0]} область',
                'district': f'{array_address[2]} район',
                'village_council': '',
                'town': f'г. {array_address[-1]}'}
    elif array_address[5] == 'аг.':
        return {'region': f'{array_address[0]} область',
                'district': f'{array_address[2]} район',
                'village_council': '',
                'town': f'г. {array_address[-1]}'}
    elif array_address[5] == 'с/с':
        return {'region': f'{array_address[0]} область',
                'district': f'{array_address[2]} район',
                'village_council': f'{array_address[4]} с/с',
                'town': f'{array_address[-2]} {array_address[-1]}'}
    elif array_address[5] == 'поселковый':
        return {'region': f'{array_address[0]} область',
                'district': f'{array_address[2]} район',
                'village_council': f'{array_address[4]} поселковый Совет',
                'town': f'{array_address[-2]} {array_address[-1]}'}
    elif array_address[-2] == 'гп':
        return {'region': f'{array_address[0]} область',
                'district': f'{array_address[2]} район',
                'village_council': '',
                'town': f'{array_address[-2]} {array_address[-1]}'}
    elif array_address[5] == 'поссовет':
        return {'region': f'{array_address[0]} область',
                'district': f'{array_address[2]} район',
                'village_council': f'{array_address[4]} поссовет с/с',
                'town': f'{array_address[-2]} {array_address[-1]}'}
    else:
        return {}


def parser_type_network(type_network):
    array_type_network = type_network.strip('\n').split(' ')
    return {'type': array_type_network[2].split('\n')[0],
            'diameter': array_type_network[3]}


def parser_coordinates(coordinates):
    return {'width': coordinates.split(' ')[0],
            'height': coordinates.split(' ')[1]}


def add_hydrants_to_db():
    wrkbk = openpyxl.load_workbook(
        '/home/fireman/work/base_of_hydrants/server/hydrant_server/main_app/utils/hydrants2.xlsx')

    sh = wrkbk.active
    result = []
    not_successful = []
    for i in range(3, sh.max_row + 1):
        if sh.cell(row=i, column=2).value == 'Пожарный гидрант':
            result.append({
                'serial_number': sh.cell(row=i, column=1).value,
                'type': sh.cell(row=i, column=2).value,
                'address': parser_address(sh.cell(row=i, column=3).value),
                'detail_address': sh.cell(row=i, column=4).value,
                'number': sh.cell(row=i, column=5).value,
                'type_network': parser_type_network(sh.cell(row=i, column=6).value),
                'technical_condition': sh.cell(row=i, column=7).value,
                'fault_tape': sh.cell(row=i, column=8).value,
                'date_last_check': sh.cell(row=i, column=9).value,
                'belonging': sh.cell(row=i, column=10).value,
                'coordinates': parser_coordinates(sh.cell(row=i, column=11).value),
                'place_plate': sh.cell(row=i, column=12).value,
                'distance_front': sh.cell(row=i, column=13).value,
                'distance_left': sh.cell(row=i, column=14).value,
                'distance_right': sh.cell(row=i, column=15).value,
                'description': sh.cell(row=i, column=16).value,
                'note': sh.cell(row=i, column=17).value,
            })
    print('File processed successfully')
    for hydrant in result:
        try:
            record = Hydrant.objects.create(serial_number=hydrant['serial_number'],
                                            number_in_map=hydrant['number'] if hydrant['number'] else '',
                                            type=hydrant['type'],
                                            address_region=hydrant['address']['region'],
                                            address_district=hydrant['address']['district'],
                                            address_village_council=hydrant['address']['village_council'],
                                            address_town=hydrant['address']['town'],
                                            address_detail=hydrant['detail_address'].strip('\n\t') if hydrant[
                                                'detail_address'] else 'None',
                                            type_of_water_supply=hydrant['type_network']['type'],
                                            diameter_drain=hydrant['type_network']['diameter'],
                                            serviceable=True if hydrant[
                                                                    'technical_condition'] == 'Исправен' else False,
                                            fault_type=hydrant['fault_tape'] if hydrant[
                                                                                    'technical_condition'] != 'Исправен' and
                                                                                hydrant['fault_tape'] else '',
                                            belonging=hydrant['belonging'] if hydrant['belonging'] else '',
                                            coordinate_width=hydrant['coordinates']['width'],
                                            coordinate_height=hydrant['coordinates']['height'],
                                            place_plate=hydrant['place_plate'] if hydrant['place_plate'] else '',
                                            distance_front=hydrant['distance_front'] if hydrant[
                                                'distance_front'] else '',
                                            distance_left=hydrant['distance_left'] if hydrant['distance_left'] else '',
                                            distance_right=hydrant['distance_right'] if hydrant[
                                                'distance_right'] else '',
                                            author=Profile.objects.first(),
                                            description=hydrant['description'] if hydrant['description'] else '',
                                            note=hydrant['note'] if hydrant['note'] else '',
                                            date_last_check=hydrant['date_last_check'],
                                            )
            print(f'{hydrant["serial_number"]}  successfully added (ERROR {len(not_successful)})')
        except KeyError:
            not_successful.append(hydrant["serial_number"])
            print(f'error {hydrant["serial_number"]}')
    if not_successful:
        print('not successful')
        print(not_successful)
        print(len(not_successful))


